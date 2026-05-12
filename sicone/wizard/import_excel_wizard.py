# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, api
from odoo.exceptions import UserError
import openpyxl
from openpyxl.utils import get_column_letter


class ImportExcelWizard(models.TransientModel):
    _name = 'sicone.import.excel.wizard'
    _description = 'Asistente para Importar Cotización de Excel'

    sale_order_id = fields.Many2one(
        'sale.order',
        string='Cotización',
        required=True,
        help='Cotización donde se cargarán los datos del Excel'
    )
    
    excel_file = fields.Binary(
        string='Archivo Excel',
        required=True,
        help='Archivo de cotización en formato Excel'
    )
    
    filename = fields.Char(
        string='Nombre del archivo',
        help='Nombre del archivo Excel cargado'
    )

    @api.model
    def default_get(self, fields_list):
        """Obtiene el sale_order del contexto si está disponible"""
        defaults = super().default_get(fields_list)
        if 'sale_order_id' in fields_list and self.env.context.get('active_id'):
            defaults['sale_order_id'] = self.env.context.get('active_id')
        return defaults

    def action_import_excel(self): 
        """
        Parsea el archivo Excel y carga los datos en la cotización (sale.order)
        """
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError('Debe seleccionar un archivo Excel')
        
        if not self.sale_order_id:
            raise UserError('Debe seleccionar una cotización')

        if self.sale_order_id.state == 'draft':
            # Limpiar líneas de la orden
            self.sale_order_id.order_line.unlink()
            # Limpiar Cotización Ponderada
            self.env['cotizacion.ponderada'].search([('sale_order_id', '=', self.sale_order_id.id)]).unlink()
            # Aquí puedes agregar limpieza de otros modelos adicionales si los creas en el futuro

        try:
            # Decodificar el archivo
            excel_data = base64.b64decode(self.excel_file)
            excel_file = io.BytesIO(excel_data)
            
            # Cargar el workbook leyendo valores calculados (no fórmulas)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active

            # --- IMPORTAR COTIZACIÓN PONDERADA ---
            if 'Cotizacion Ponderada' in wb.sheetnames:
                ws_ponderada = wb['Cotizacion Ponderada']
                CotizacionPonderada = self.env['cotizacion.ponderada']
                # Eliminar registros previos para esta cotización
                CotizacionPonderada.search([('sale_order_id', '=', self.sale_order_id.id)]).unlink()
                # Buscar la fila de títulos
                start_row = None
                for idx, row in enumerate(ws_ponderada.iter_rows(values_only=True)):
                    if row and 'CONCEPTO' in [str(cell).upper() for cell in row]:
                        start_row = idx
                        break
                if start_row is not None:
                    # Leer las filas de datos
                    for row in ws_ponderada.iter_rows(min_row=start_row+2, values_only=True):
                        concepto = row[0]
                        valor = row[1]
                        peso = row[2]
                        if concepto and valor is not None and peso is not None:
                            CotizacionPonderada.create({
                                'sale_order_id': self.sale_order_id.id,
                                'concepto': str(concepto),
                                'valor': float(valor) if isinstance(valor, (int, float)) else 0.0,
                                'peso': float(peso) if isinstance(peso, (int, float)) else 0.0,
                            })

            # Parsear la estructura del Excel
            parsed_data = self._parse_excel_structure(ws)
            # Cargar los datos en la cotización (DISEÑOS Y PLANIFICACIÓN)
            self._load_data_to_quotation(parsed_data)

            # Cargar la sección ESPECIFICACIONES GENERALES DEL PROYECTO
            self._import_especificaciones_generales(ws)

            # Guardar el archivo como adjunto a la cotización
            from datetime import datetime
            now_str = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            base_name = self.filename or 'cotizacion_importada.xlsx'
            if '.' in base_name:
                name_part, ext = base_name.rsplit('.', 1)
                new_name = f"{name_part}_{now_str}.{ext}"
            else:
                new_name = f"{base_name}_{now_str}"
            self.env['ir.attachment'].create({
                'name': new_name,
                'datas': self.excel_file,
                'res_model': 'sale.order',
                'res_id': self.sale_order_id.id,
                'type': 'binary',
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Importación exitosa',
                    'message': f'Se cargaron {len(parsed_data)} líneas en la cotización (más especificaciones generales) y el archivo fue adjuntado.',
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        
        except Exception as e:
            raise UserError(f'Error al procesar el Excel: {str(e)}')

    def _import_especificaciones_generales(self, ws):
        """
        Importa la sección 'ESPECIFICACIONES GENERALES DEL PROYECTO' (ejemplo: mampostería)
        Agrega una línea de sección y líneas de ítem con valores correctos.
        """
        SaleOrderLine = self.env['sale.order.line']
        sequence = SaleOrderLine.search_count([('order_id', '=', self.sale_order_id.id)]) + 100

        # Buscar la fila donde inicia la sección
        start_row = None
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_text = ' '.join([str(v).upper() for v in row if v is not None])
            if 'ESPECIFICACIONES GENERALES DEL PROYECTO' in row_text:
                start_row = idx
                break

        if start_row is None:
            # No se encontró la sección, no hacer nada
            return

        # Crear línea de sección
        SaleOrderLine.create({
            'order_id': self.sale_order_id.id,
            'display_type': 'line_section',
            'name': 'ESPECIFICACIONES GENERALES DEL PROYECTO',
            'sequence': sequence,
        })
        sequence += 1

        # Buscar la fila de valores (la fila después de la fila de títulos)
        # start_row es el índice de la fila con "ESPECIFICACIONES GENERALES DEL PROYECTO"
        # Las filas de títulos están en start_row+1, los valores en start_row+2
        value_row = None
        try:
            # Leer la fila directamente usando el número de fila real (start_row es índice 0-based)
            row_num = start_row + 3  # +1 para fila, +2 para saltar a fila de valores
            row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
            if row and any(v is not None for v in row):
                value_row = row
        except Exception as e:
            pass
        
        if not value_row:
            return

        # Extraer valores según columnas del Excel:
        # Fila 24: [0]empty, [1]m², [2]845.00, [3]$67.000, [4]$56.615.000, [5]$7.500, [6]$6.337.500, [7]$45.000, [8]$38.025.000, [9]$100.977.500
        descripcion = 'Mampostería'
        uom = value_row[1] if len(value_row) > 1 else ''
        cantidad = value_row[2] if len(value_row) > 2 and isinstance(value_row[2], (int, float)) else 1.0

        def parse_val(val):
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                cleaned = val.replace('$','').replace(' ','').replace('.','').replace(',','.')
                try:
                    return float(cleaned)
                except Exception:
                    return 0.0
            return 0.0

        # Posiciones correctas: [4]Materiales, [6]Equipos, [8]Mano_Obra, [9]Subtotal
        materiales = parse_val(value_row[4]) if len(value_row) > 4 else 0.0
        equipos = parse_val(value_row[6]) if len(value_row) > 6 else 0.0
        mano_obra = parse_val(value_row[8]) if len(value_row) > 8 else 0.0
        subtotal = parse_val(value_row[9]) if len(value_row) > 9 else 0.0

        # Buscar producto por nombre o referencia
        product = self._get_product_by_name('MAMPOSTERIA')
        if not product:
            raise UserError("No se encontró el producto 'MAMPOSTERIA' en Odoo para la sección de especificaciones.")

        SaleOrderLine.create({
            'order_id': self.sale_order_id.id,
            'product_id': product.id,
            'name': descripcion,
            'product_uom_qty': cantidad,
            'price_unit': subtotal/ cantidad if cantidad else 0.0,
            'materiales': materiales,
            'equipos': equipos,
            'mano_obra': mano_obra,
            'sequence': sequence,
        })
        sequence += 1

        # Agregar ítem Estructura General: tomar siempre la fila 27 (índice 27, 1-based)
        try:
            value_row_estructura = list(ws.iter_rows(min_row=27, max_row=27, values_only=True))[0]
            if value_row_estructura and any(v is not None for v in value_row_estructura):
                descripcion_estr = 'Estructura General'
                ref_estr = 'ESTRUCTURA_GENERAL'
                uom_estr = value_row_estructura[1] if len(value_row_estructura) > 1 else ''
                cantidad_estr = value_row_estructura[2] if len(value_row_estructura) > 2 and isinstance(value_row_estructura[2], (int, float)) else 1.0
                materiales_estr = parse_val(value_row_estructura[4]) if len(value_row_estructura) > 4 else 0.0
                equipos_estr = parse_val(value_row_estructura[6]) if len(value_row_estructura) > 6 else 0.0
                mano_obra_estr = parse_val(value_row_estructura[8]) if len(value_row_estructura) > 8 else 0.0
                subtotal_estr = parse_val(value_row_estructura[9]) if len(value_row_estructura) > 9 else 0.0

                product_estr = self._get_product_by_name(ref_estr)
                if not product_estr:
                    raise UserError("No se encontró el producto 'ESTRUCTURA_GENERAL' en Odoo para la sección de especificaciones.")

                SaleOrderLine.create({
                    'order_id': self.sale_order_id.id,
                    'product_id': product_estr.id,
                    'name': descripcion_estr,
                    'product_uom_qty': cantidad_estr,
                    'price_unit': subtotal_estr / cantidad_estr if cantidad_estr else 0.0,
                    'materiales': materiales_estr,
                    'equipos': equipos_estr,
                    'mano_obra': mano_obra_estr,
                    'sequence': sequence,
                })
                sequence += 1
        except Exception:
            pass

        # Agregar ítem Cubierta, Superboard y Manto: tomar siempre la fila 30 (índice 30, 1-based)
        try:
            value_row_cubierta = list(ws.iter_rows(min_row=30, max_row=30, values_only=True))[0]
            if value_row_cubierta and any(v is not None for v in value_row_cubierta):
                descripcion_cub = 'Cubierta, Superboard y Manto'
                ref_cub = 'CUBIERTA'
                uom_cub = value_row_cubierta[1] if len(value_row_cubierta) > 1 else ''
                cantidad_cub = value_row_cubierta[2] if len(value_row_cubierta) > 2 and isinstance(value_row_cubierta[2], (int, float)) else 1.0
                materiales_cub = parse_val(value_row_cubierta[4]) if len(value_row_cubierta) > 4 else 0.0
                equipos_cub = parse_val(value_row_cubierta[6]) if len(value_row_cubierta) > 6 else 0.0
                mano_obra_cub = parse_val(value_row_cubierta[8]) if len(value_row_cubierta) > 8 else 0.0
                subtotal_cub = parse_val(value_row_cubierta[9]) if len(value_row_cubierta) > 9 else 0.0

                product_cub = self._get_product_by_name(ref_cub)
                if not product_cub:
                    raise UserError("No se encontró el producto 'CUBIERTA' en Odoo para la sección de especificaciones.")

                SaleOrderLine.create({
                    'order_id': self.sale_order_id.id,
                    'product_id': product_cub.id,
                    'name': descripcion_cub,
                    'product_uom_qty': cantidad_cub,
                    'price_unit': subtotal_cub / cantidad_cub if cantidad_cub else 0.0,
                    'materiales': materiales_cub,
                    'equipos': equipos_cub,
                    'mano_obra': mano_obra_cub,
                    'sequence': sequence,
                })
                sequence += 1
        except Exception:
            pass

        # Agregar ítem Cubierta, Superboard y Shingle: tomar siempre la fila 43 (índice 43, 1-based)
        try:
            value_row_shingle = list(ws.iter_rows(min_row=43, max_row=43, values_only=True))[0]
            if value_row_shingle and any(v is not None for v in value_row_shingle):
                descripcion_shingle = 'Cubierta, Superboard y Shingle'
                ref_shingle = 'CUBIERTA_SHINGLE'
                uom_shingle = value_row_shingle[1] if len(value_row_shingle) > 1 else ''
                cantidad_shingle = value_row_shingle[2] if len(value_row_shingle) > 2 and isinstance(value_row_shingle[2], (int, float)) else 1.0
                materiales_shingle = parse_val(value_row_shingle[4]) if len(value_row_shingle) > 3 else 0.0
                equipos_shingle = parse_val(value_row_shingle[6]) if len(value_row_shingle) > 5 else 0.0
                mano_obra_shingle = parse_val(value_row_shingle[8]) if len(value_row_shingle) > 7 else 0.0
                subtotal_shingle = parse_val(value_row_shingle[9]) if len(value_row_shingle) > 9 else 0.0

                product_shingle = self._get_product_by_name(ref_shingle)
                if not product_shingle:
                    raise UserError("No se encontró el producto 'CUBIERTA_SHINGLE' en Odoo para la sección de especificaciones.")

                SaleOrderLine.create({
                    'order_id': self.sale_order_id.id,
                    'product_id': product_shingle.id,
                    'name': descripcion_shingle,
                    'product_uom_qty': cantidad_shingle,
                    'price_unit': subtotal_shingle / cantidad_shingle if cantidad_shingle else 0.0,
                    'materiales': materiales_shingle,
                    'equipos': equipos_shingle,
                    'mano_obra': mano_obra_shingle,
                    'sequence': sequence,
                })
                sequence += 1
        except Exception:
            pass

        # Agregar ítem Entrepiso Placa Fácil: tomar siempre la fila 46 (índice 46, 1-based)
        try:
            value_row_entrepiso = list(ws.iter_rows(min_row=46, max_row=46, values_only=True))[0]
            if value_row_entrepiso and any(v is not None for v in value_row_entrepiso):
                descripcion_entrepiso = 'Entrepiso Placa Fácil'
                ref_entrepiso = 'ENTREPISO_PLACA'
                uom_entrepiso = value_row_entrepiso[1] if len(value_row_entrepiso) > 1 else ''
                cantidad_entrepiso = value_row_entrepiso[2] if len(value_row_entrepiso) > 2 and isinstance(value_row_entrepiso[2], (int, float)) else 1.0
                materiales_entrepiso = parse_val(value_row_entrepiso[4]) if len(value_row_entrepiso) > 3 else 0.0
                equipos_entrepiso = parse_val(value_row_entrepiso[6]) if len(value_row_entrepiso) > 5 else 0.0
                mano_obra_entrepiso = parse_val(value_row_entrepiso[8]) if len(value_row_entrepiso) > 7 else 0.0
                subtotal_entrepiso = parse_val(value_row_entrepiso[9]) if len(value_row_entrepiso) > 9 else 0.0

                product_entrepiso = self._get_product_by_name(ref_entrepiso)
                if not product_entrepiso:
                    raise UserError("No se encontró el producto 'ENTREPISO_PLACA' en Odoo para la sección de especificaciones.")

                SaleOrderLine.create({
                    'order_id': self.sale_order_id.id,
                    'product_id': product_entrepiso.id,
                    'name': descripcion_entrepiso,
                    'product_uom_qty': cantidad_entrepiso,
                    'price_unit': subtotal_entrepiso / cantidad_entrepiso if cantidad_entrepiso else 0.0,
                    'materiales': materiales_entrepiso,
                    'equipos': equipos_entrepiso,
                    'mano_obra': mano_obra_entrepiso,
                    'sequence': sequence,
                })
                sequence += 1
        except Exception:
            pass

        # Agregar ítem Pérgolas y Estructura sin Techo: tomar siempre la fila 52 (índice 52, 1-based)
        try:
            value_row_pergolas = list(ws.iter_rows(min_row=52, max_row=52, values_only=True))[0]
            if value_row_pergolas and any(v is not None for v in value_row_pergolas):
                descripcion_pergolas = 'Pérgolas y Estructura sin Techo'
                ref_pergolas = 'PERGOLAS_SIN_TECHO'
                uom_pergolas = value_row_pergolas[1] if len(value_row_pergolas) > 1 else ''
                cantidad_pergolas = value_row_pergolas[2] if len(value_row_pergolas) > 2 and isinstance(value_row_pergolas[2], (int, float)) else 1.0
                materiales_pergolas = parse_val(value_row_pergolas[4]) if len(value_row_pergolas) > 4 else 0.0
                equipos_pergolas = parse_val(value_row_pergolas[6]) if len(value_row_pergolas) > 6 else 0.0
                mano_obra_pergolas = parse_val(value_row_pergolas[8]) if len(value_row_pergolas) > 8 else 0.0
                subtotal_pergolas = parse_val(value_row_pergolas[9]) if len(value_row_pergolas) > 9 else 0.0

                product_pergolas = self._get_product_by_name(ref_pergolas)
                if not product_pergolas:
                    raise UserError("No se encontró el producto 'PERGOLAS_SIN_TECHO' en Odoo para la sección de especificaciones.")

                SaleOrderLine.create({
                    'order_id': self.sale_order_id.id,
                    'product_id': product_pergolas.id,
                    'name': descripcion_pergolas,
                    'product_uom_qty': cantidad_pergolas,
                    'price_unit': subtotal_pergolas / cantidad_pergolas if cantidad_pergolas else 0.0,
                    'materiales': materiales_pergolas,
                    'equipos': equipos_pergolas,
                    'mano_obra': mano_obra_pergolas,
                    'sequence': sequence,
                })
                sequence += 1
        except Exception:
            pass

        # Guardar resumen de costos directos en campos de solo lectura de sale.order
        try:
            row_estructura = list(ws.iter_rows(min_row=58, max_row=58, values_only=True))[0]
            row_otros = list(ws.iter_rows(min_row=59, max_row=59, values_only=True))[0]
            row_disenos = list(ws.iter_rows(min_row=60, max_row=60, values_only=True))[0]
            row_total = list(ws.iter_rows(min_row=61, max_row=61, values_only=True))[0]
            estructura_val = 0.0
            otros_val = 0.0
            disenos_val = 0.0
            total_val = 0.0
            def parse_val(val):
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    cleaned = val.replace('$','').replace(' ','').replace('.','').replace(',','.')
                    try:
                        return float(cleaned)
                    except Exception:
                        return 0.0
                return 0.0
            # Column J is index 9
            if row_estructura and len(row_estructura) > 9:
                estructura_val = parse_val(row_estructura[9])
            if row_otros and len(row_otros) > 9:
                otros_val = parse_val(row_otros[9])
            if row_disenos and len(row_disenos) > 9:
                disenos_val = parse_val(row_disenos[9])
            if row_total and len(row_total) > 9:
                total_val = parse_val(row_total[9])
            self.sale_order_id.write({
                'resumen_estructura': estructura_val,
                'resumen_otros': otros_val,
                'resumen_disenos': disenos_val,
                'resumen_total_costo_directo': total_val,
            })
        except Exception:
            pass

        # Guardar ADMINISTRACIÓN DEL PROYECTO en campos de solo lectura de sale.order
        try:
            row_comision_ventas = list(ws.iter_rows(min_row=64, max_row=64, values_only=True))[0]
            row_imprevistos = list(ws.iter_rows(min_row=65, max_row=65, values_only=True))[0]
            row_pct_admon = list(ws.iter_rows(min_row=66, max_row=66, values_only=True))[0]
            row_vlr_sugerido_admon = list(ws.iter_rows(min_row=67, max_row=67, values_only=True))[0]
            row_calculo_admon = list(ws.iter_rows(min_row=68, max_row=68, values_only=True))[0]
            row_administracion = list(ws.iter_rows(min_row=69, max_row=69, values_only=True))[0]
            row_logistica = list(ws.iter_rows(min_row=70, max_row=70, values_only=True))[0]
            row_pct_utilidad = list(ws.iter_rows(min_row=71, max_row=71, values_only=True))[0]
            row_vlr_sugerido_utilidad = list(ws.iter_rows(min_row=72, max_row=72, values_only=True))[0]
            row_calculo_utilidad = list(ws.iter_rows(min_row=73, max_row=73, values_only=True))[0]
            row_utilidad = list(ws.iter_rows(min_row=74, max_row=74, values_only=True))[0]
            row_total_aiu = list(ws.iter_rows(min_row=75, max_row=75, values_only=True))[0]

            def parse_pct(val):
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    cleaned = val.replace('%','').replace(',','.').replace(' ','')
                    try:
                        return float(cleaned)
                    except Exception:
                        return 0.0
                return 0.0

            vals_admin = {
                'comision_ventas': parse_val(row_comision_ventas[9]) if row_comision_ventas and len(row_comision_ventas) > 9 else 0.0,
                'imprevistos': parse_val(row_imprevistos[9]) if row_imprevistos and len(row_imprevistos) > 9 else 0.0,
                'porcentaje_sugerido_admon': parse_pct(row_pct_admon[9]) if row_pct_admon and len(row_pct_admon) > 9 else 0.0,
                'vlr_sugerido_admon': parse_val(row_vlr_sugerido_admon[9]) if row_vlr_sugerido_admon and len(row_vlr_sugerido_admon) > 9 else 0.0,
                'calculo_admon': parse_val(row_calculo_admon[9]) if row_calculo_admon and len(row_calculo_admon) > 9 else 0.0,
                'administracion': parse_val(row_administracion[9]) if row_administracion and len(row_administracion) > 9 else 0.0,
                'logistica': parse_val(row_logistica[9]) if row_logistica and len(row_logistica) > 9 else 0.0,
                'porcentaje_sugerido_utilidad': parse_pct(row_pct_utilidad[9]) if row_pct_utilidad and len(row_pct_utilidad) > 9 else 0.0,
                'vlr_sugerido_utilidad': parse_val(row_vlr_sugerido_utilidad[9]) if row_vlr_sugerido_utilidad and len(row_vlr_sugerido_utilidad) > 9 else 0.0,
                'calculo_utilidad': parse_val(row_calculo_utilidad[9]) if row_calculo_utilidad and len(row_calculo_utilidad) > 9 else 0.0,
                'utilidad': parse_val(row_utilidad[9]) if row_utilidad and len(row_utilidad) > 9 else 0.0,
                'total_aiu': parse_val(row_total_aiu[9]) if row_total_aiu and len(row_total_aiu) > 9 else 0.0,
            }
            self.sale_order_id.write(vals_admin)
        except Exception:
            pass

        # Guardar VALOR TOTAL DEL PROYECTO en campos de solo lectura de sale.order
        try:
            row_directos = list(ws.iter_rows(min_row=78, max_row=78, values_only=True))[0]
            row_aiu = list(ws.iter_rows(min_row=79, max_row=79, values_only=True))[0]
            row_subtotal = list(ws.iter_rows(min_row=80, max_row=80, values_only=True))[0]
            row_descuento = list(ws.iter_rows(min_row=81, max_row=81, values_only=True))[0]
            row_total_costo = list(ws.iter_rows(min_row=82, max_row=82, values_only=True))[0]
            row_valor_m2 = list(ws.iter_rows(min_row=83, max_row=83, values_only=True))[0]

            def parse_val_m2(val):
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    cleaned = val.replace('$','').replace(' ','').replace('.','').replace(',','.').replace('*10M','').replace('m2','').replace('²','')
                    try:
                        return float(cleaned)
                    except Exception:
                        return 0.0
                return 0.0

            vals_total = {
                'valor_total_directos': parse_val(row_directos[9]) if row_directos and len(row_directos) > 9 else 0.0,
                'valor_total_aiu': parse_val(row_aiu[9]) if row_aiu and len(row_aiu) > 9 else 0.0,
                'valor_total_subtotal_costo_directo': parse_val(row_subtotal[9]) if row_subtotal and len(row_subtotal) > 9 else 0.0,
                'valor_total_descuento': parse_val(row_descuento[9]) if row_descuento and len(row_descuento) > 9 else 0.0,
                'valor_total_total_costo_directo': parse_val(row_total_costo[9]) if row_total_costo and len(row_total_costo) > 9 else 0.0,
                'valor_total_m2': parse_val_m2(row_valor_m2[9]) if row_valor_m2 and len(row_valor_m2) > 9 else 0.0,
            }
            self.sale_order_id.write(vals_total)
        except Exception:
            pass

                # Guardar VALOR TOTAL DEL PROYECTO CON CIMENTACIONES Y COMPLEMENTARIOS en campos de solo lectura de sale.order
        try:
            row_costo_directo = list(ws.iter_rows(min_row=86, max_row=86, values_only=True))[0]
            row_cimentaciones = list(ws.iter_rows(min_row=87, max_row=87, values_only=True))[0]
            row_complementarios = list(ws.iter_rows(min_row=88, max_row=88, values_only=True))[0]
            row_total_costo = list(ws.iter_rows(min_row=89, max_row=89, values_only=True))[0]
            row_valor_m2 = list(ws.iter_rows(min_row=90, max_row=90, values_only=True))[0]

            def parse_val_m2_cim_comp(val):
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    cleaned = val.replace('$','').replace(' ','').replace('.','').replace(',','.').replace('*10M','').replace('m2','').replace('²','')
                    try:
                        return float(cleaned)
                    except Exception:
                        return 0.0
                return 0.0

            vals_cim_comp = {
                'costo_directo_cim_comp': parse_val(row_costo_directo[9]) if row_costo_directo and len(row_costo_directo) > 9 else 0.0,
                'cimentaciones': parse_val(row_cimentaciones[9]) if row_cimentaciones and len(row_cimentaciones) > 9 else 0.0,
                'complementarios': parse_val(row_complementarios[9]) if row_complementarios and len(row_complementarios) > 9 else 0.0,
                'total_costo_proyecto_cim_comp': parse_val(row_total_costo[9]) if row_total_costo and len(row_total_costo) > 9 else 0.0,
                'valor_m2_cim_comp': parse_val_m2_cim_comp(row_valor_m2[9]) if row_valor_m2 and len(row_valor_m2) > 9 else 0.0,
            }
            self.sale_order_id.write(vals_cim_comp)
        except Exception:
            pass

        # Agregar capítulo de CARGOS Y AJUSTES (AIU y Descuentos)
        try:
            SaleOrderLine = self.env['sale.order.line']
            sequence = SaleOrderLine.search_count([('order_id', '=', self.sale_order_id.id)]) + 1000
            # Crear línea de sección
            SaleOrderLine.create({
                'order_id': self.sale_order_id.id,
                'display_type': 'line_section',
                'name': 'CARGOS Y AJUSTES',
                'sequence': sequence,
            })
            sequence += 1
            # Obtener valores de AIU (J75) y Descuentos (J81)
            row_aiu = list(ws.iter_rows(min_row=75, max_row=75, values_only=True))[0]
            row_desc = list(ws.iter_rows(min_row=81, max_row=81, values_only=True))[0]
            def parse_val(val):
                if isinstance(val, (int, float)):
                    return float(val)
                if isinstance(val, str):
                    cleaned = val.replace('$','').replace(' ','').replace('.','').replace(',','.')
                    try:
                        return float(cleaned)
                    except Exception:
                        return 0.0
                return 0.0
            aiu_val = parse_val(row_aiu[9]) if row_aiu and len(row_aiu) > 9 else 0.0
            desc_val = parse_val(row_desc[9]) if row_desc and len(row_desc) > 9 else 0.0
            # Buscar productos por referencia
            product_aiu = self._get_product_by_name('AIU')
            product_desc = self._get_product_by_name('DESCUENTOS')
            if product_aiu:
                SaleOrderLine.create({
                    'order_id': self.sale_order_id.id,
                    'product_id': product_aiu.id,
                    'name': product_aiu.display_name,
                    'product_uom_qty': 1.0,
                    'price_unit': aiu_val,
                    'sequence': sequence,
                    'capitulo': 'CARGOS Y AJUSTES',
                })
                sequence += 1
            if product_desc:
                SaleOrderLine.create({
                    'order_id': self.sale_order_id.id,
                    'product_id': product_desc.id,
                    'name': product_desc.display_name,
                    'product_uom_qty': 1.0,
                    'price_unit': desc_val,
                    'sequence': sequence,
                    'capitulo': 'CARGOS Y AJUSTES',
                })
        except Exception:
            pass

    def _parse_excel_structure(self, ws):
        """
        Parsea la estructura del Excel según el formato F-CDO-2014
        Retorna una lista de diccionarios con la información extraída
        """
        parsed_data = []
        
        # Recorrer las filas del Excel
        for row in ws.iter_rows(min_row=1, values_only=False):
            # Obtener valores de celdas
            row_data = [cell.value for cell in row]
            
            # Filtrar filas vacías
            if all(v is None for v in row_data):
                continue
            
            parsed_data.append(row_data)
        
        # Aquí procesaremos la estructura específica del Excel
        return self._extract_chapters_and_items(parsed_data, ws)

    def _extract_chapters_and_items(self, raw_data, ws=None):
        """
        Extrae los capítulos e ítems de los datos crudos del Excel
        Basado en la estructura F-CDO-2014
        Los ítems están en columnas horizontales, no filas verticales
        Lee los valores de celdas fijas según la plantilla.
        """
        # Mapeo: nombre en Excel, nombre en Odoo, índice de columna (basado en la plantilla real)
        # B=1, D=3, F=5, H=7 (índices 0-based)
        items_map = [
            {'excel_name': 'DISEÑO ARQUITECTONICO', 'odoo_name': 'DISEÑO_ARQ', 'col': 1},  # B
            {'excel_name': 'DISEÑO ESTRUCTURAL', 'odoo_name': 'DISEÑO_ESTR', 'col': 3},    # D
            {'excel_name': 'DESARROLLO DEL PROYECTO', 'odoo_name': 'DESARROLLO_PROY', 'col': 5}, # F
            {'excel_name': 'VISITA TECNICA', 'odoo_name': 'VISITATECNICA', 'col': 7},     # H
        ]

        # Buscar la fila de títulos y la fila de valores
        title_row_idx = None
        value_row_idx = None
        for idx, row in enumerate(raw_data):
            row_text = ' '.join([str(v).upper() for v in row if v is not None])
            if 'DISEÑO ARQUITECTONICO' in row_text and 'DISEÑO ESTRUCTURAL' in row_text:
                title_row_idx = idx
                value_row_idx = idx + 1
                break

        if title_row_idx is None or value_row_idx is None:
            raise UserError(
                'No se encontró la sección "DISEÑOS Y PLANIFICACIÓN" en el archivo Excel. '
                'Verifique que el archivo tenga la estructura correcta.'
            )

        value_row = raw_data[value_row_idx]
        items = []
        def parse_number(val):
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                # Eliminar símbolos de moneda, puntos de miles, espacios y convertir coma decimal a punto
                cleaned = val.replace('$', '').replace(' ', '').replace('.', '').replace(',', '.')
                try:
                    return float(cleaned)
                except Exception:
                    return 0.0
            return 0.0

        # Leer cantidad de la celda J9 (área construida)
        cantidad = 1.0
        if ws is not None:
            try:
                cantidad = parse_number(ws['J9'].value)
            except Exception:
                cantidad = 1.0

        for item_def in items_map:
            col = item_def['col']
            subtotal = 0.0
            if col < len(value_row):
                cell_val = value_row[col]
                subtotal = parse_number(cell_val)
            items.append({
                'excel_name': item_def['excel_name'],
                'odoo_name': item_def['odoo_name'],
                'description': item_def['excel_name'],
                'quantity': cantidad,
                'subtotal': subtotal,
            })

        # Ya no se valida que haya al menos un ítem con valor
        return items

    def _extract_item_from_row(self, row):
        """
        Extrae la información de un ítem de una fila del Excel
        Estructura: cantidad, materiales, equipos, mano de obra, subtotal
        """
        if len(row) < 5:
            return None
        
        # Intentar extraer información numérica
        numeric_values = []
        for i, cell in enumerate(row):
            if isinstance(cell, (int, float)):
                numeric_values.append({
                    'column': i,
                    'value': cell
                })
        
        if not numeric_values:
            return None
        
        # Por ahora, creamos un ítem básico
        return {
            'description': str(row[0]) if row[0] else 'Ítem sin descripción',
            'values': numeric_values
        }

    def _extract_disenos_item(self, row, item_name):
        """
        Extrae un ítem de la sección DISEÑOS Y PLANIFICACIÓN
        Estructura típica: nombre | valor1 | valor2 | valor3 | valor4 | subtotal
        """
        # Los valores están en las columnas, ignorando la primera (nombre)
        values = []
        for cell in row[1:]:
            if isinstance(cell, (int, float)):
                values.append(cell)
        
        # Si hay valores numéricos, crear el ítem
        if values:
            # El subtotal es el último valor
            subtotal = values[-1] if values else 0
            
            return {
                'name': item_name,
                'description': item_name,
                'quantity': 1.0,
                'subtotal': subtotal,
                'all_values': values  # Guardar todos los valores para debug
            }
        
        return None
    

    def _load_data_to_quotation(self, items):
        """
        Carga los ítems extraídos en la cotización (sale.order)
        """
        if not items:
            raise UserError('No se encontraron ítems en la sección DISEÑOS Y PLANIFICACIÓN del archivo')
        
        SaleOrderLine = self.env['sale.order.line']
        sequence = 1
        
        # Crear línea de sección para DISEÑOS Y PLANIFICACIÓN
        SaleOrderLine.create({
            'order_id': self.sale_order_id.id,
            'display_type': 'line_section',
            'name': 'DISEÑOS Y PLANIFICACIÓN',
            'sequence': sequence,
        })
        sequence += 1
        
        # Crear líneas para cada ítem
        for item in items:
            # Buscar producto usando el nombre mapeado de Odoo
            product = self._get_product_by_name(item['odoo_name'])
            
            if not product:
                raise UserError(f"No se encontró el producto '{item['odoo_name']}' en Odoo. "
                               f"Nombre en Excel: {item['excel_name']}")
            
            SaleOrderLine.create({
                'order_id': self.sale_order_id.id,
                'product_id': product.id,
                'name': product.display_name,
                'product_uom_qty': item.get('quantity', 1.0),
                'price_unit': item.get('subtotal', 0.0),
                'sequence': sequence,
            })
            sequence += 1

    def _get_or_create_product(self, item):
        """
        Obtiene o crea un producto basado en la descripción del ítem
        """
        Product = self.env['product.product']
        
        # Buscar producto existente por nombre
        products = Product.search([
            ('name', 'ilike', item.get('description', ''))
        ], limit=1)
        
        if products:
            return products[0]
        
        # Crear nuevo producto
        return Product.create({
            'name': item.get('description', 'Producto importado'),
            'type': 'service',
            'list_price': 0.0,
        })

    def _get_product_by_name(self, product_identifier):
        """
        Busca un producto por:
        1. Nombre exacto (case-insensitive)
        2. Referencia/código
        3. Nombre parcial
        
        product_identifier puede ser: nombre completo, código corto, etc.
        """
        Product = self.env['product.product']
        
        # Primero intentar búsqueda por referencia/código
        products = Product.search([
            ('default_code', '=', product_identifier)
        ], limit=1)
        
        if products:
            return products[0]
        
        # Búsqueda case-insensitive por nombre exacto
        products = Product.search([
            ('name', '=ilike', product_identifier)
        ], limit=1)
        
        if products:
            return products[0]
        
        # Búsqueda parcial por nombre
        products = Product.search([
            ('name', 'ilike', product_identifier)
        ], limit=1)
        
        if products:
            return products[0]
        
        # Búsqueda parcial por referencia
        products = Product.search([
            ('default_code', 'ilike', product_identifier)
        ], limit=1)
        
        return products[0] if products else None
