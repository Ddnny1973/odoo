# -*- coding: utf-8 -*-
"""
WIZARD: Importación del Libro Auxiliar Contable
================================================
Importa movimientos contables desde el archivo Excel anual
generado por el sistema contable.

LÓGICA DE CARGA INTELIGENTE:
- Detecta el rango de fechas del archivo nuevo
- Compara con datos ya cargados (zona de solapamiento = último mes)
- Actualiza movimientos modificados en la zona de solapamiento
- Carga solo datos nuevos fuera de esa zona
- Nunca duplica movimientos (clave única: fecha+comprobante+secuencia+cuenta+CC)

CLASIFICACIÓN DE CUENTAS:
- 41XXXXX + crédito → Ingreso de Proyecto
- 71/72/73XXXXX + débito → Egreso (Materiales/MO/Variables/Admin)
- 51XXXXX + CC=ADMINISTRATIVO → Gasto Fijo Administrativo
- 1110050102 + crédito → Egreso real de caja (cuenta bancaria)

Para registrar datos históricos de migración use:
⚠️ Carga Histórica Inicial (Solo usar una vez)
"""
import base64
import io
from datetime import datetime, date
from collections import defaultdict
from odoo import models, fields, api
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class WizardLibroAuxiliar(models.TransientModel):
    _name = 'sicone.wizard.libro.auxiliar'
    _description = 'Importar Libro Auxiliar Contable'

    archivo_excel = fields.Binary(
        string='Archivo Excel',
        required=True,
        help='Libro auxiliar contable exportado del sistema contable (formato xlsx)'
    )
    archivo_nombre = fields.Char(string='Nombre del archivo')

    # ── Preview ───────────────────────────────────────────────────────
    fecha_min_archivo = fields.Date(string='Fecha inicial del archivo', readonly=True)
    fecha_max_archivo = fields.Date(string='Fecha final del archivo', readonly=True)
    fecha_max_cargado = fields.Date(string='Último dato cargado', readonly=True)
    total_filas = fields.Integer(string='Total filas detectadas', readonly=True)
    filas_nuevas = fields.Integer(string='Filas nuevas', readonly=True)
    filas_solapamiento = fields.Integer(string='Filas en zona de solapamiento', readonly=True)
    cc_sin_mapeo = fields.Text(string='Centros de costo sin mapeo', readonly=True)
    cuentas_sin_clasificar = fields.Text(string='Cuentas sin clasificar', readonly=True)
    advertencia = fields.Text(string='Advertencia', readonly=True)

    company_id = fields.Many2one(
        'res.company', default=lambda self: self.env.company
    )
    currency_id = fields.Many2one(
        'res.currency', related='company_id.currency_id', readonly=True
    )

    @api.onchange('archivo_excel')
    def _onchange_archivo_excel(self):
        if not self.archivo_excel:
            return
        try:
            filas = self._leer_filas()
            if not filas:
                self.advertencia = 'No se encontraron filas válidas en el archivo.'
                return

            fechas = [f['fecha'] for f in filas if f.get('fecha')]
            self.fecha_min_archivo = min(fechas) if fechas else False
            self.fecha_max_archivo = max(fechas) if fechas else False
            self.total_filas = len(filas)

            # Verificar último dato cargado en BD
            ultimo = self.env['sicone.movimiento.contable'].search(
                [('fuente', '=', 'libro_auxiliar')],
                order='fecha desc', limit=1
            )
            self.fecha_max_cargado = ultimo.fecha if ultimo else False

            # Calcular zona de solapamiento (último mes cargado)
            if self.fecha_max_cargado:
                from dateutil.relativedelta import relativedelta
                inicio_solap = self.fecha_max_cargado - relativedelta(months=1)
                nuevas = [f for f in filas if f['fecha'] > self.fecha_max_cargado]
                solap = [f for f in filas if inicio_solap <= f['fecha'] <= self.fecha_max_cargado]
                self.filas_nuevas = len(nuevas)
                self.filas_solapamiento = len(solap)
            else:
                self.filas_nuevas = len(filas)
                self.filas_solapamiento = 0

            # Detectar CC sin mapeo
            mapeo_cc = {cc.nombre_cc: cc.proyecto_id
                        for cc in self.env['sicone.centro.costo'].search([])}
            cc_archivo = set(f['cc'] for f in filas if f.get('cc'))
            sin_mapeo = cc_archivo - set(mapeo_cc.keys())
            # Filtrar CC que no son de proyectos (ADMINISTRATIVO, COMERCIAL, etc.)
            cc_no_proyecto = {'ADMINISTRATIVO -', 'COMERCIAL -', 'PLANTA -',
                              'ALAMMO -', 'BODEGAS CADI -', 'ESPIRAL -',
                              'FABRICASAS -', 'JUAN ESTEBAN LINCE -'}
            sin_mapeo_proyectos = sin_mapeo - cc_no_proyecto
            if sin_mapeo_proyectos:
                self.cc_sin_mapeo = '\n'.join(sorted(sin_mapeo_proyectos))

            # Detectar cuentas sin clasificar (solo 7XXXXX)
            clasificacion = {c.nombre_cuenta: c.categoria
                             for c in self.env['sicone.cuenta.clasificacion'].search([])}
            cuentas_egreso = set(f['nombre_cuenta'] for f in filas
                                 if f.get('nombre_cuenta') and
                                 str(f.get('codigo', '')).startswith('7'))
            sin_clas = cuentas_egreso - set(clasificacion.keys())
            if sin_clas:
                self.cuentas_sin_clasificar = '\n'.join(sorted(sin_clas))

        except Exception as e:
            self.advertencia = f'Error al analizar archivo: {str(e)}'

    def _leer_filas(self):
        """Lee y filtra las filas válidas del libro auxiliar"""
        if not openpyxl:
            raise UserError('openpyxl no está instalado.')

        contenido = base64.b64decode(self.archivo_excel)
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)

        # Buscar Sheet1 o primera hoja
        hoja = 'Sheet1' if 'Sheet1' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[hoja]

        filas = []
        for i, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
            codigo = row[0]
            if not codigo:
                continue
            codigo_str = str(codigo).strip()

            # Ignorar filas de subtotales/encabezados
            if not codigo_str.replace('.', '').isdigit():
                continue

            # Solo procesar cuentas relevantes
            prefix = codigo_str[:2]
            if prefix not in ('41', '71', '72', '73', '51', '53', '11'):
                continue

            nombre_cuenta = row[1]
            fecha_raw = row[7]
            cc = row[10]
            debito = row[13]
            credito = row[14]
            comprobante = row[5]
            secuencia = str(row[6]) if row[6] else ''

            # Parsear fecha
            fecha = None
            if fecha_raw:
                try:
                    if isinstance(fecha_raw, (date, datetime)):
                        fecha = fecha_raw.date() if isinstance(fecha_raw, datetime) else fecha_raw
                    else:
                        fecha = datetime.strptime(str(fecha_raw), '%d/%m/%Y').date()
                except Exception:
                    continue

            if not fecha:
                continue

            # Determinar tipo y monto
            tipo = None
            monto = 0.0
            if prefix == '41' and credito:
                tipo = 'ingreso'
                monto = float(credito)
            elif prefix in ('71', '72', '73') and debito:
                tipo = 'egreso'
                monto = float(debito)
            elif prefix == '51' and debito:
                tipo = 'egreso'
                monto = float(debito)
            elif prefix == '53' and debito:
                tipo = 'egreso'
                monto = float(debito)
            elif codigo_str == '1110050102' and credito:
                # Cuenta bancaria — créditos son egresos reales de caja
                tipo = 'egreso'
                monto = float(credito)

            if not tipo or monto <= 0:
                continue

            filas.append({
                'codigo': codigo_str,
                'nombre_cuenta': str(nombre_cuenta) if nombre_cuenta else '',
                'identificacion': str(row[2]) if row[2] else '',
                'nombre_tercero': str(row[4]) if row[4] else '',
                'comprobante': str(comprobante) if comprobante else '',
                'secuencia': secuencia,
                'fecha': fecha,
                'descripcion': str(row[8]) if row[8] else '',
                'detalle': str(row[9]) if row[9] else '',
                'cc': str(cc).strip() if cc else '',
                'debito': float(debito) if debito else 0.0,
                'credito': float(credito) if credito else 0.0,
                'tipo': tipo,
                'monto': monto,
                'prefix': prefix,
            })

        return filas

    def action_importar(self):
        """Ejecuta la importación del libro auxiliar"""
        self.ensure_one()

        filas = self._leer_filas()
        if not filas:
            raise UserError('No se encontraron movimientos válidos en el archivo.')

        # Cargar tablas de referencia
        mapeo_cc = {cc.nombre_cc: cc.proyecto_id
                    for cc in self.env['sicone.centro.costo'].search([])}
        clasificacion = {c.nombre_cuenta: c.categoria
                         for c in self.env['sicone.cuenta.clasificacion'].search([])}

        # Zona de solapamiento — último mes ya cargado
        ultimo = self.env['sicone.movimiento.contable'].search(
            [('fuente', '=', 'libro_auxiliar')],
            order='fecha desc', limit=1
        )
        fecha_corte = ultimo.fecha if ultimo else None

        contadores = defaultdict(int)
        pagos_distribuidos = 0

        for fila in filas:
            # Determinar si es zona de solapamiento
            es_solapamiento = fecha_corte and fila['fecha'] <= fecha_corte

            # Buscar movimiento existente
            existente = self.env['sicone.movimiento.contable'].search([
                ('fecha', '=', fila['fecha']),
                ('comprobante', '=', fila['comprobante']),
                ('secuencia', '=', fila['secuencia']),
                ('codigo_cuenta', '=', fila['codigo']),
                ('nombre_cc', '=', fila['cc']),
            ], limit=1)

            if existente and not es_solapamiento:
                contadores['omitidos'] += 1
                continue

            # Determinar proyecto desde centro de costo
            proyecto = mapeo_cc.get(fila['cc'])

            # Determinar categoría
            categoria = self._determinar_categoria(
                fila['prefix'], fila['codigo'], fila['nombre_cuenta'],
                fila['cc'], clasificacion
            )

            vals = {
                'fecha': fila['fecha'],
                'comprobante': fila['comprobante'],
                'secuencia': fila['secuencia'],
                'codigo_cuenta': fila['codigo'],
                'nombre_cuenta': fila['nombre_cuenta'],
                'identificacion': fila['identificacion'],
                'nombre_tercero': fila['nombre_tercero'],
                'descripcion': fila['descripcion'],
                'detalle': fila['detalle'],
                'nombre_cc': fila['cc'],
                'proyecto_id': proyecto.id if proyecto else False,
                'categoria': categoria,
                'tipo': fila['tipo'],
                'debito': fila['debito'],
                'credito': fila['credito'],
                'monto': fila['monto'],
                'fuente': 'banco' if fila['codigo'] == '1110050102' else 'libro_auxiliar',
            }

            if existente and es_solapamiento and not existente.clasificacion_manual:
                existente.write(vals)
                contadores['actualizados'] += 1
            else:
                movimiento = self.env['sicone.movimiento.contable'].create(vals)
                contadores['creados'] += 1

                # Si es ingreso de proyecto → distribuir contra hitos
                if fila['tipo'] == 'ingreso' and proyecto:
                    pagos_dist = self._distribuir_ingreso(movimiento, proyecto)
                    pagos_distribuidos += pagos_dist

        # ── Recalcular Cash Flow de todos los proyectos afectados ────
        proyectos_afectados = self.env['sicone.proyecto'].search([
            ('id', 'in', self.env['sicone.movimiento.contable'].search([
                ('fuente', '=', 'libro_auxiliar'),
                ('proyecto_id', '!=', False),
            ]).mapped('proyecto_id').ids)
        ])
        for proyecto in proyectos_afectados:
            proyecto.action_recalcular_cashflow()

        mensaje = (
            f"Importación completada:\n"
            f"• {contadores['creados']} movimientos nuevos\n"
            f"• {contadores['actualizados']} actualizados (solapamiento)\n"
            f"• {contadores['omitidos']} omitidos (ya existían)\n"
            f"• {pagos_distribuidos} pagos distribuidos en hitos\n"
            f"• {len(proyectos_afectados)} proyectos con Cash Flow actualizado"
        )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': '✅ Importación exitosa',
                'message': mensaje,
                'type': 'success',
                'sticky': True, 'next': {'type': 'ir.actions.act_window_close'},
            }
        }

    # CC corporativos — no corresponden a un proyecto específico
    CC_GASTOS_FIJOS = {
        'ADMINISTRATIVO -', 'COMERCIAL -', 'PLANTA -',
        'ALAMMO -', 'BODEGAS CADI -', 'ESPIRAL -',
        'FABRICASAS -', 'JUAN ESTEBAN LINCE -', '',
    }

    def _determinar_categoria(self, prefix, codigo_cuenta, nombre_cuenta, cc, clasificacion):
        """
        Determina la categoría del movimiento.
        - 41x crédito        -> ingreso_proyecto
        - 1110050102 crédito -> egreso real de caja (categoría según CC)
        - 7x débito          -> busca en tabla clasificacion
        - 51x/53x            -> si CC corporativo -> gasto_fijo
                                si CC de proyecto -> clasificacion o 'administracion'
        """
        if prefix == '41':
            return 'ingreso_proyecto'
        if codigo_cuenta == '1110050102':
            # Egreso bancario directo — categoría según CC
            cc_limpio = (cc or '').strip()
            if cc_limpio in self.CC_GASTOS_FIJOS:
                return 'gasto_fijo'
            cat = clasificacion.get(nombre_cuenta)
            return cat if cat else 'sin_clasificar'
        if prefix in ('51', '53'):
            cc_limpio = (cc or '').strip()
            if cc_limpio in self.CC_GASTOS_FIJOS:
                return 'gasto_fijo'
            cat = clasificacion.get(nombre_cuenta)
            return cat if cat else 'administracion'
        cat = clasificacion.get(nombre_cuenta)
        if cat:
            return cat
        return 'sin_clasificar'

    def _distribuir_ingreso(self, movimiento, proyecto):
        """Distribuye un ingreso contra los hitos pendientes del proyecto"""
        hitos = proyecto.hito_ids.filtered(
            lambda h: h.estado in ('pendiente', 'parcial')
        ).sorted('secuencia')

        if not hitos:
            return 0

        monto_restante = movimiento.monto
        pagos_creados = 0

        for hito in hitos:
            if monto_restante <= 0:
                break
            saldo = hito.monto - hito.cobrado
            if saldo <= 0:
                continue
            aplicar = min(monto_restante, saldo)
            monto_restante -= aplicar

            self.env['sicone.pago'].create({
                'hito_id': hito.id,
                'fecha': movimiento.fecha,
                'monto': aplicar,
                'referencia': movimiento.comprobante or 'Libro Auxiliar',
                'notas': f"Importado desde libro auxiliar — {movimiento.descripcion}",
            })
            pagos_creados += 1

            movimiento.write({'hito_id': hito.id})

        return pagos_creados
