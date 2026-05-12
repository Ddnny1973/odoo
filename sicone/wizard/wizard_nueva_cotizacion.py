# -*- coding: utf-8 -*-
from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import io

class WizardNuevaCotizacion(models.TransientModel):
    _name = 'sicone.wizard.nueva.cotizacion'
    _description = 'Wizard Nueva Cotización'

    cliente = fields.Char(string='Cliente')
    direccion = fields.Char(string='Dirección')
    area_base = fields.Float(string='Área Base (m²)', digits=(12, 2))
    archivo_excel = fields.Binary(string='Archivo Excel de Cotización')
    archivo_nombre = fields.Char(string='Nombre del Archivo')
    modo = fields.Selection([
        ('excel', 'Cargar desde Excel'),
        ('manual', 'Crear vacía'),
    ], string='Modo', default='excel', required=True)

    @api.onchange('modo')
    def _onchange_modo(self):
        if self.modo == 'manual':
            self.archivo_excel = False
            self.archivo_nombre = False

    def action_test_lectura(self):
        """Muestra exactamente qué está leyendo el parser del Excel"""
        self.ensure_one()
        if not self.archivo_excel:
            raise UserError('Seleccione un archivo Excel primero.')
        datos = self._parsear_excel()
        
        msg = "=== LECTURA EXCEL ===\n\n"
        msg += f"Cliente: {datos.get('cliente','')}\n"
        msg += f"Dirección: {datos.get('direccion','')}\n"
        msg += f"Área base: {datos.get('area_base',0):,.2f}\n\n"
        
        msg += "--- DISEÑOS ---\n"
        for d in datos.get('disenos', []):
            msg += f"  {d['nombre']}: {d['precio_unitario']:,.0f}/m²\n"
        
        msg += "\n--- ESTRUCTURA ---\n"
        for e in datos.get('estructura', []):
            msg += f"  {e['nombre']}: mat={e['precio_materiales']:,.0f} eq={e['precio_equipos']:,.0f} mo={e['precio_mano_obra']:,.0f}\n"
        
        msg += "\n--- MAMPOSTERÍA ---\n"
        for m in datos.get('mamposteria', []):
            msg += f"  cant={m['cantidad']:,.2f} mat={m['precio_materiales']:,.0f}\n"
        
        msg += "\n--- TECHOS ---\n"
        for t in datos.get('techos', []):
            msg += f"  {t['nombre']}: cant={t['cantidad']:,.2f} mat={t['precio_materiales']:,.0f}\n"

        aiu = datos.get('aiu_cap1', {})
        msg += f"\n--- AIU CAP1 ---\n"
        msg += f"  Comisión: {aiu.get('cap1_comision_ventas',0):,.0f}\n"
        msg += f"  Imprevistos: {aiu.get('cap1_imprevistos',0):,.0f}\n"
        msg += f"  Administración: {aiu.get('cap1_administracion',0):,.0f}\n"
        msg += f"  Logística: {aiu.get('cap1_logistica',0):,.0f}\n"
        msg += f"  Utilidad: {aiu.get('cap1_utilidad',0):,.0f}\n"

        msg += f"\n--- CIMENTACIÓN ---\n"
        msg += f"  Opción activa: {datos.get('cimentacion_opcion_activa','')}\n"
        msg += f"  Op1 logística: {datos.get('cim_op1_logistica',0):,.0f}\n"
        msg += f"  Op2 logística: {datos.get('cim_op2_logistica',0):,.0f}\n"
        for c in datos.get('cimentacion_op1', []):
            msg += f"  Op1 {c['nombre']}: cant={c['cantidad']:,.2f} precio={c['precio_unitario']:,.0f}\n"
        for c in datos.get('cimentacion_op2', []):
            msg += f"  Op2 {c['nombre']}: cant={c['cantidad']:,.2f} precio={c['precio_unitario']:,.0f}\n"

        msg += f"\n--- COMPLEMENTARIOS ---\n"
        msg += f"  comp_pct_comision: {datos.get('comp_pct_comision',0):,.2f}%\n"
        msg += f"  comp_pct_aiu: {datos.get('comp_pct_aiu',0):,.2f}%\n"
        msg += f"  comp_logistica: {datos.get('comp_logistica',0):,.0f}\n"
        for c in datos.get('complementarios', []):
            msg += f"  {c['nombre']}: cant={c['cantidad']:,.2f} precio={c['precio_unitario']:,.0f}\n"

        raise UserError(msg)

    def action_crear_cotizacion(self):
        self.ensure_one()

        vals_cotizacion = {
            'cliente': self.cliente or '',
            'direccion': self.direccion or '',
        }
        area_base = self.area_base or 0.0
        datos_version = {}

        if self.modo == 'excel' and self.archivo_excel:
            try:
                datos = self._parsear_excel()
                if datos.get('cliente') and not vals_cotizacion['cliente']:
                    vals_cotizacion['cliente'] = datos.pop('cliente')
                else:
                    datos.pop('cliente', None)
                if datos.get('direccion') and not vals_cotizacion['direccion']:
                    vals_cotizacion['direccion'] = datos.pop('direccion')
                else:
                    datos.pop('direccion', None)
                if datos.get('area_base'):
                    area_base = datos.pop('area_base')
                else:
                    datos.pop('area_base', None)
                datos_version = datos
            except Exception as e:
                raise UserError(f'Error al leer el Excel: {str(e)}')
        elif self.modo == 'excel' and not self.archivo_excel:
            raise UserError('Debe seleccionar un archivo Excel o cambiar a modo manual.')

        # Si viene de una cotización existente, la actualiza en lugar de crear nueva
        cotizacion_id = self.env.context.get('default_cotizacion_id')
        if cotizacion_id:
            cotizacion = self.env['sicone.cotizacion'].browse(cotizacion_id)
            cotizacion.write(vals_cotizacion)
        else:
            cotizacion = self.env['sicone.cotizacion'].create(vals_cotizacion)

        if cotizacion.version_ids:
            version = cotizacion.version_ids[0]
            # Extraer áreas desde techos
            area_cubierta = 0.0
            area_entrepiso = 0.0
            manto = 0.0
            shingle = 0.0
            for t in datos_version.get('techos', []):
                if t.get('nombre') == 'cubierta_manto':
                    manto = t.get('cantidad', 0.0)
                elif t.get('nombre') == 'cubierta_shingle':
                    shingle = t.get('cantidad', 0.0)
                elif t.get('nombre') == 'entrepiso':
                    area_entrepiso = t.get('cantidad', 0.0)
            area_cubierta = max(manto, shingle)

            version.write({
                'area_base': area_base,
                'area_cubierta': area_cubierta,
                'area_entrepiso': area_entrepiso,
            })
            if datos_version:
                self._poblar_version(version, datos_version)
                # Forzar recompute via SQL directo
                version._compute_capitulo2()
                version._compute_capitulo1()
                version._compute_totales_generales()
                version.env.cr.execute(
                    "UPDATE sicone_cotizacion_version SET "
                    "total_cim_op1=%s, total_cim_op2=%s, "
                    "subtotal_complementarios=%s, total_complementarios=%s, "
                    "total_capitulo2=%s, total_general=%s WHERE id=%s",
                    (version.total_cim_op1, version.total_cim_op2,
                     version.subtotal_complementarios, version.total_complementarios,
                     version.total_capitulo2, version.total_general, version.id)
                )
                cotizacion._compute_totales_consolidados()
                version.env.cr.execute(
                    "UPDATE sicone_cotizacion SET "
                    "total_capitulo1=%s, total_capitulo2=%s, total_general=%s WHERE id=%s",
                    (cotizacion.total_capitulo1, cotizacion.total_capitulo2,
                     cotizacion.total_general, cotizacion.id)
                )

        return {
            'name': 'Cotización',
            'type': 'ir.actions.act_window',
            'res_model': 'sicone.cotizacion',
            'view_mode': 'form',
            'res_id': cotizacion.id,
            'target': 'current',
        }

    def _get_val(self, ws, row, col):
        """Lee un valor numérico de la hoja, retorna 0.0 si es None"""
        v = ws.cell(row=row, column=col).value
        if v is None:
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    def _parsear_excel(self):
        import openpyxl
        contenido = base64.b64decode(self.archivo_excel)
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        datos = {}

        # ===================== HOJA COTIZACIÓN =====================
        if 'Cotización' in wb.sheetnames:
            ws = wb['Cotización']

            # Encabezado
            cliente = ws.cell(row=5, column=4).value
            if cliente:
                datos['cliente'] = str(cliente).strip()
            direccion = ws.cell(row=6, column=4).value
            if direccion:
                datos['direccion'] = str(direccion).strip()
            datos['area_base'] = self._get_val(ws, 15, 4)

            # --- DISEÑOS (F20: C2=arq precio, C3=arq subtotal, C4=estr precio, C5=estr subtotal...) ---
            # Siempre leer subtotal (Cn+1) y dividir por area para obtener precio_unitario
            # Esto funciona tanto si C4 es precio/m² como si es subtotal directo
            area = datos.get('area_base') or 1.0
            def precio_diseno(subtotal):
                return round(subtotal / area, 2) if area else 0.0

            datos['disenos'] = [
                {'nombre': 'arq',  'precio_unitario': precio_diseno(self._get_val(ws, 20, 3))},
                {'nombre': 'estr', 'precio_unitario': precio_diseno(self._get_val(ws, 20, 5))},
                {'nombre': 'dev',  'precio_unitario': precio_diseno(self._get_val(ws, 20, 7))},
                {'nombre': 'vis',  'precio_unitario': precio_diseno(self._get_val(ws, 20, 9))},
            ]

            # --- MAMPOSTERÍA (F24) ---
            datos['mamposteria'] = [{
                'nombre': 'Mampostería',
                'unidad': 'm2',
                'cantidad': self._get_val(ws, 24, 3),
                'precio_materiales': self._get_val(ws, 24, 4),
                'precio_equipos': self._get_val(ws, 24, 6),
                'precio_mano_obra': self._get_val(ws, 24, 8),
            }]

            # --- ESTRUCTURA (F27) - lee C10 (subtotal final) para evitar discrepancias de fórmula ---
            datos['estructura'] = [{
                'nombre': 'Estructura General',
                'unidad': 'gl',
                'cantidad': 1,
                'precio_materiales': self._get_val(ws, 27, 5),
                'precio_equipos': self._get_val(ws, 27, 7),
                'precio_mano_obra': self._get_val(ws, 27, 10) - self._get_val(ws, 27, 5) - self._get_val(ws, 27, 7),
            }]

            # --- TECHOS (filas con datos) ---
            # cubierta_manto=F30, cubierta_shingle=F43, entrepiso=F46, pergolas=F52
            datos['techos'] = [
                {
                    'nombre': 'cubierta_manto',
                    'unidad': 'm2',
                    'cantidad': self._get_val(ws, 30, 3),
                    'precio_materiales': self._get_val(ws, 30, 4),
                    'precio_equipos': self._get_val(ws, 30, 6),
                    'precio_mano_obra': self._get_val(ws, 30, 8),
                },
                {
                    'nombre': 'cubierta_shingle',
                    'unidad': 'm2',
                    'cantidad': self._get_val(ws, 43, 3),
                    'precio_materiales': self._get_val(ws, 43, 4),
                    'precio_equipos': self._get_val(ws, 43, 6),
                    'precio_mano_obra': self._get_val(ws, 43, 8),
                },
                {
                    'nombre': 'entrepiso',
                    'unidad': 'm2',
                    'cantidad': self._get_val(ws, 46, 3),
                    'precio_materiales': self._get_val(ws, 46, 4),
                    'precio_equipos': self._get_val(ws, 46, 6),
                    'precio_mano_obra': self._get_val(ws, 46, 8),
                },
                {
                    'nombre': 'pergolas',
                    'unidad': 'm2',
                    'cantidad': self._get_val(ws, 52, 3),
                    'precio_materiales': self._get_val(ws, 52, 4),
                    'precio_equipos': self._get_val(ws, 52, 6),
                    'precio_mano_obra': self._get_val(ws, 52, 8),
                },
            ]

            # --- AIU CAPÍTULO 1 ---
            datos['aiu_cap1'] = {
                'cap1_comision_ventas': self._get_val(ws, 64, 10),
                'cap1_imprevistos': self._get_val(ws, 65, 10),
                'cap1_pct_administracion': self._get_val(ws, 66, 10) * 100,
                'cap1_administracion': self._get_val(ws, 69, 10),
                'cap1_logistica': self._get_val(ws, 70, 10),
                'cap1_pct_utilidad': self._get_val(ws, 71, 10) * 100,
                'cap1_utilidad': self._get_val(ws, 74, 10),
            }

        # ===================== HOJA ADMINISTRACIÓN =====================
        if 'Administracion' in wb.sheetnames:
            ws_adm = wb['Administracion']

            # Personal Profesional F23-F28
            prof_map = {
                'dir_obra': 23, 'sup_tecnico': 24, 'prof_presupuesto': 25,
                'arq_disenador': 26, 'oficial_obra': 27, 'ayudante_obra': 28,
            }
            datos['personal_prof'] = []
            for key, fila in prof_map.items():
                datos['personal_prof'].append({
                    'tipo': 'profesional',
                    'nombre': key,
                    'cantidad': int(self._get_val(ws_adm, fila, 5) or 1),
                    'valor_mes': self._get_val(ws_adm, fila, 6),
                    'pct_prestaciones': self._get_val(ws_adm, fila, 7) * 100,
                    'dedicacion': self._get_val(ws_adm, fila, 8),
                    'meses': int(self._get_val(ws_adm, fila, 9) or 0),
                })

            # Personal Administrativo F33-F39
            admin_map = {
                'prof_procesos': 33, 'gerente_gral': 34, 'compras': 35,
                'contabilidad': 36, 'atencion_cliente': 37,
                'mantenimiento': 38, 'gestion_humana': 39,
            }
            datos['personal_admin'] = []
            for key, fila in admin_map.items():
                datos['personal_admin'].append({
                    'tipo': 'administrativo',
                    'nombre': key,
                    'cantidad': int(self._get_val(ws_adm, fila, 5) or 1),
                    'valor_mes': self._get_val(ws_adm, fila, 6),
                    'pct_prestaciones': self._get_val(ws_adm, fila, 7) * 100,
                    'dedicacion': self._get_val(ws_adm, fila, 8),
                    'meses': int(self._get_val(ws_adm, fila, 9) or 0),
                })

        # ===================== HOJA CIMENTACIONES Y COMPLEMENTARIOS =====================
        if 'Cimentaciones y Complementarios' in wb.sheetnames:
            ws_cim = wb['Cimentaciones y Complementarios']

            # Opción 1: Pilas (F4) + Vigas y Losa (F5)
            datos['cimentacion_op1'] = [
                {
                    'opcion': 'opcion1', 'nombre': 'pilas', 'unidad': 'und',
                    'cantidad': self._get_val(ws_cim, 4, 7),
                    'precio_unitario': self._get_val(ws_cim, 4, 8),
                },
                {
                    'opcion': 'opcion1', 'nombre': 'vigas_losa', 'unidad': 'm2',
                    'cantidad': self._get_val(ws_cim, 5, 7),
                    'precio_unitario': self._get_val(ws_cim, 5, 8),
                },
            ]
            datos['cim_op1_pct_comision'] = self._get_val(ws_cim, 7, 7) * 100
            datos['cim_op1_pct_aiu'] = self._get_val(ws_cim, 8, 7) * 100
            datos['cim_op1_logistica'] = self._get_val(ws_cim, 9, 9)

            # Opción 2: Pilotes (F14) + Vigas y Losa (F15)
            datos['cimentacion_op2'] = [
                {
                    'opcion': 'opcion2', 'nombre': 'pilotes', 'unidad': 'und',
                    'cantidad': self._get_val(ws_cim, 14, 7),
                    'precio_unitario': self._get_val(ws_cim, 14, 8),
                },
                {
                    'opcion': 'opcion2', 'nombre': 'vigas_losa', 'unidad': 'm2',
                    'cantidad': self._get_val(ws_cim, 15, 7),
                    'precio_unitario': self._get_val(ws_cim, 15, 8),
                },
            ]
            datos['cim_op2_pct_comision'] = self._get_val(ws_cim, 17, 7) * 100
            datos['cim_op2_pct_aiu'] = self._get_val(ws_cim, 18, 7) * 100
            datos['cim_op2_logistica'] = self._get_val(ws_cim, 19, 9)

            # Complementarios F24-F33
            comp_map = [
                (24, 'aguas_lluvias'), (25, 'hidrosanitaria'), (26, 'escalas'),
                (27, 'campamento'), (28, 'cerramiento'), (29, 'canoa_metalica'),
                (30, 'ruana_metalica'), (31, 'revoque'), (32, 'fajas_ranuras'),
                (33, 'otros'),
            ]
            datos['complementarios'] = []
            for fila, key in comp_map:
                cant = self._get_val(ws_cim, fila, 7)
                precio = self._get_val(ws_cim, fila, 8)
                if cant or precio:
                    datos['complementarios'].append({
                        'nombre': key,
                        'unidad': 'und',
                        'cantidad': cant,
                        'precio_unitario': precio,
                    })
            # Opción activa de cimentación (F3 C11)
            opcion_activa_txt = ws_cim.cell(row=3, column=11).value
            if opcion_activa_txt and 'Opción 1' in str(opcion_activa_txt):
                datos['cimentacion_opcion_activa'] = 'opcion1'
            elif opcion_activa_txt and 'Opción 2' in str(opcion_activa_txt):
                datos['cimentacion_opcion_activa'] = 'opcion2'

            datos['comp_pct_comision'] = self._get_val(ws_cim, 35, 7) * 100
            datos['comp_pct_aiu'] = self._get_val(ws_cim, 36, 7) * 100
            datos['comp_logistica'] = self._get_val(ws_cim, 37, 9)

        return datos

    def _poblar_version(self, version, datos):
        """Crea los registros de detalle en la versión"""
        env = self.env

        # Escribir primero todos los porcentajes y logísticas
        # para que el compute tenga los valores correctos desde el inicio
        aiu = datos.get('aiu_cap1', {})
        version.write({
            'cimentacion_opcion_activa': datos.get('cimentacion_opcion_activa', False),
            'cim_op1_pct_comision': datos.get('cim_op1_pct_comision', 3.0),
            'cim_op1_pct_aiu': datos.get('cim_op1_pct_aiu', 47.0),
            'cim_op1_logistica': datos.get('cim_op1_logistica', 0.0),
            'cim_op2_pct_comision': datos.get('cim_op2_pct_comision', 3.0),
            'cim_op2_pct_aiu': datos.get('cim_op2_pct_aiu', 47.0),
            'cim_op2_logistica': datos.get('cim_op2_logistica', 0.0),
            'comp_pct_comision': datos.get('comp_pct_comision', 0.0),
            'comp_pct_aiu': datos.get('comp_pct_aiu', 0.0),
            'comp_logistica': datos.get('comp_logistica', 0.0),
            **aiu,
        })
        version.env.flush_all()

        # Diseños
        for d in datos.get('disenos', []):
            if d['precio_unitario']:
                env['sicone.cotizacion.diseno'].create({
                    'version_id': version.id,
                    'nombre': d['nombre'],
                    'precio_unitario': d['precio_unitario'],
                })

        # Mampostería
        for m in datos.get('mamposteria', []):
            env['sicone.cotizacion.mamposteria'].create({
                'version_id': version.id,
                'nombre': m['nombre'],
                'unidad': m.get('unidad', 'm2'),
                'cantidad': m['cantidad'],
                'precio_materiales': m['precio_materiales'],
                'precio_equipos': m['precio_equipos'],
                'precio_mano_obra': m['precio_mano_obra'],
            })

        # Estructura
        for e in datos.get('estructura', []):
            env['sicone.cotizacion.estructura'].create({
                'version_id': version.id,
                'nombre': e['nombre'],
                'unidad': e.get('unidad', 'gl'),
                'cantidad': e['cantidad'],
                'precio_materiales': e['precio_materiales'],
                'precio_equipos': e['precio_equipos'],
                'precio_mano_obra': e['precio_mano_obra'],
            })

        # Techos
        for t in datos.get('techos', []):
            env['sicone.cotizacion.techo'].create({
                'version_id': version.id,
                'nombre': t['nombre'],
                'unidad': t.get('unidad', 'm2'),
                'cantidad': t['cantidad'],
                'precio_materiales': t['precio_materiales'],
                'precio_equipos': t['precio_equipos'],
                'precio_mano_obra': t['precio_mano_obra'],
            })

        # Personal Profesional
        for p in datos.get('personal_prof', []):
            if p['valor_mes']:
                env['sicone.cotizacion.personal'].create({
                    'version_id': version.id,
                    **p,
                })

        # Personal Administrativo
        for p in datos.get('personal_admin', []):
            if p['valor_mes']:
                env['sicone.cotizacion.personal'].create({
                    'version_id': version.id,
                    **p,
                })

        # Cimentaciones Op1
        for c in datos.get('cimentacion_op1', []):
            if c['cantidad'] or c['precio_unitario']:
                env['sicone.cotizacion.cimentacion'].create({
                    'version_id': version.id, **c,
                })
        # Cimentación Opción 2
        for c in datos.get('cimentacion_op2', []):
            if c['cantidad'] or c['precio_unitario']:
                env['sicone.cotizacion.cimentacion'].create({
                    'version_id': version.id, **c,
                })
        # Complementarios
        for c in datos.get('complementarios', []):
            env['sicone.cotizacion.complementario'].create({
                'version_id': version.id, **c,
            })

    def _forzar_recompute(self, version):
        """Fuerza recompute correcto de campos stored en Odoo 18"""
        campos = [
            'total_complementarios', 'total_cim_op1', 'total_cim_op2',
            'total_capitulo2', 'subtotal_cim_op1', 'subtotal_cim_op2',
            'subtotal_complementarios'
        ]
        for campo in campos:
            version.env.add_to_compute(
                version._fields[campo], version
            )
        version.flush_recordset(campos)
        # Propagar a cotizacion
        cotizacion = version.cotizacion_id
        for campo in ['total_capitulo1', 'total_capitulo2', 'total_general']:
            cotizacion.env.add_to_compute(
                cotizacion._fields[campo], cotizacion
            )
        cotizacion.flush_recordset(['total_capitulo1', 'total_capitulo2', 'total_general'])
